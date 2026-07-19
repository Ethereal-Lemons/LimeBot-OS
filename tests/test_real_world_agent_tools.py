import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace


class TestCalculatorTool(unittest.IsolatedAsyncioTestCase):
    async def test_calculates_cloud_monthly_and_annual_cost(self):
        from core.bus import MessageBus
        from core.tools import Toolbox

        toolbox = Toolbox(
            allowed_paths=[str(Path.cwd())],
            bus=MessageBus(),
            config=SimpleNamespace(skills=SimpleNamespace(enabled=[])),
        )

        self.assertEqual(await toolbox.calculate("0.0104*730"), "Result: 7.592")
        self.assertEqual(await toolbox.calculate("0.0104*730*12"), "Result: 91.104")

    async def test_rejects_code_and_unbounded_power(self):
        from core.bus import MessageBus
        from core.tools import Toolbox

        toolbox = Toolbox(
            allowed_paths=[str(Path.cwd())],
            bus=MessageBus(),
            config=SimpleNamespace(skills=SimpleNamespace(enabled=[])),
        )

        self.assertTrue((await toolbox.calculate("__import__('os')")).startswith("Error:"))
        self.assertTrue((await toolbox.calculate("2**1000")).startswith("Error:"))


class TestSpreadsheetTool(unittest.IsolatedAsyncioTestCase):
    async def test_creates_styled_formula_workbook(self):
        from openpyxl import load_workbook
        from core.bus import MessageBus
        from core.tools import Toolbox

        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "cloud_costs.xlsx"
            toolbox = Toolbox(
                allowed_paths=[temp_dir],
                bus=MessageBus(),
                config=SimpleNamespace(skills=SimpleNamespace(enabled=[])),
            )
            result = await toolbox.create_spreadsheet(
                str(target),
                [
                    {
                        "name": "Comparison",
                        "rows": [
                            ["Provider", "Hourly USD", "Monthly Cost USD", "Annual Total USD"],
                            ["Azure B1s", 0.0104, "=B2*730", "=C2*12"],
                        ],
                    },
                    {
                        "name": "Sources",
                        "rows": [
                            ["Source", "URL"],
                            ["Azure API", "https://prices.azure.com/api/retail/prices"],
                        ],
                    },
                ],
                title="Cloud price comparison",
            )

            self.assertTrue(result.startswith("Successfully created spreadsheet"))
            self.assertTrue(target.exists())
            workbook = load_workbook(target, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Comparison", "Sources"])
            self.assertEqual(workbook["Comparison"]["C2"].value, "=B2*730")
            self.assertEqual(workbook["Comparison"].freeze_panes, "A2")
            self.assertEqual(workbook["Sources"]["B2"].hyperlink.target, "https://prices.azure.com/api/retail/prices")
            self.assertEqual(workbook["Comparison"]["A1"].font.bold, True)

    async def test_rejects_non_xlsx_and_oversized_workbook(self):
        from core.bus import MessageBus
        from core.tools import Toolbox

        with tempfile.TemporaryDirectory() as temp_dir:
            toolbox = Toolbox(
                allowed_paths=[temp_dir],
                bus=MessageBus(),
                config=SimpleNamespace(skills=SimpleNamespace(enabled=[])),
            )
            bad_extension = await toolbox.create_spreadsheet(
                str(Path(temp_dir) / "costs.csv"), [{"name": "Data", "rows": [["x"]]}]
            )
            too_many_sheets = await toolbox.create_spreadsheet(
                str(Path(temp_dir) / "costs.xlsx"),
                [{"name": f"S{i}", "rows": [["x"]]} for i in range(21)],
            )

            self.assertTrue(bad_extension.startswith("Error:"))
            self.assertTrue(too_many_sheets.startswith("Error:"))


class TestAgentCompletionReserve(unittest.TestCase):
    def test_artifact_intent_and_tool_filter_preserve_delivery_tools(self):
        from core.loop import AgentLoop

        definitions = [
            {"function": {"name": "web_search"}},
            {"function": {"name": "browser_navigate"}},
            {"function": {"name": "browser_download"}},
            {"function": {"name": "create_spreadsheet"}},
            {"function": {"name": "calculate"}},
            {"function": {"name": "send_media"}},
        ]

        self.assertTrue(AgentLoop._artifact_delivery_requested("compare prices and send the Excel"))
        reserved = AgentLoop._artifact_reserve_tool_definitions(definitions)
        names = [item["function"]["name"] for item in reserved]
        self.assertEqual(names, ["create_spreadsheet", "calculate", "send_media"])


class TestCommandPreflight(unittest.TestCase):
    def test_rejects_windows_incompatible_chaining_before_execution(self):
        from core.bus import MessageBus
        from core.tools import Toolbox

        toolbox = Toolbox(
            allowed_paths=[str(Path.cwd())],
            bus=MessageBus(),
            config=SimpleNamespace(
                skills=SimpleNamespace(enabled=[]), allow_unsafe_commands=False
            ),
        )

        self.assertIn("forbidden", toolbox.validate_command("mkdir out && python -V").lower())
        self.assertIn("forbidden", toolbox.validate_command("python - <<'PY'").lower())
        self.assertIsNone(toolbox.validate_command("python -V"))
